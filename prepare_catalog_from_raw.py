#!/usr/bin/env python3
"""
Prepara un Excel "crudo" (como el de la captura: Producto, Precio Unidad, Precio Mayor, Condición)
para dejarlo con las columnas estándar que usa catalog_import.py.

Entrada esperada (hoja activa):
  - Producto
  - Precio Unidad
  - Precio Mayor (puede venir vacío => sin mayorista)
  - Condición (umbral de mayorista, ej. "DESDE 10 UNIDADES")

Salida: catalog_prepared.xlsx con columnas:
  sku, nombre, categoria, precio_unit, precio_mayor, umbral_mayor, favorito, superfavorito, visible

Reglas:
  - sku se genera automáticamente a partir de nombre (slug ascii). Si hay colisiones, se les agrega sufijo -2, -3...
  - categoria se deja "sin_categoria" para que puedas editar luego.
  - favorito=0, superfavorito=0, visible=1 por defecto.
  - precio_mayor vacío -> None (sin mayorista).
"""
from __future__ import annotations

import unicodedata
from pathlib import Path
from typing import Dict, Tuple

from openpyxl import load_workbook, Workbook


def slugify(text: str) -> str:
    normalized = unicodedata.normalize("NFKD", text)
    ascii_text = normalized.encode("ascii", "ignore").decode("ascii")
    result = []
    for ch in ascii_text.lower():
        if ch.isalnum():
            result.append(ch)
        elif ch in {" ", "-", "_", "."}:
            result.append("-")
    slug = "".join(result)
    while "--" in slug:
        slug = slug.replace("--", "-")
    return slug.strip("-") or "producto"


def unique_slug(base: str, existing: Dict[str, int]) -> str:
    if base not in existing:
        existing[base] = 1
        return base
    counter = existing[base] + 1
    while f"{base}-{counter}" in existing:
        counter += 1
    existing[base] = counter
    return f"{base}-{counter}"


def read_raw(path: Path):
    wb = load_workbook(path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("El archivo está vacío.")
    header = [str(h).strip().lower() for h in rows[0]]
    try:
        idx_producto = header.index("producto")
        idx_pu = header.index("precio unidad")
        idx_pm = header.index("precio mayor")
        idx_cond = header.index("condición") if "condición" in header else header.index("condicion")
    except ValueError as exc:
        raise ValueError("Faltan columnas: se necesitan Producto, Precio Unidad, Precio Mayor, Condición") from exc
    data = []
    for raw in rows[1:]:
        if raw is None or all(cell is None or str(cell).strip() == "" for cell in raw):
            continue
        nombre = str(raw[idx_producto] or "").strip()
        if not nombre:
            continue
        precio_unit = raw[idx_pu]
        precio_mayor = raw[idx_pm]
        condicion = str(raw[idx_cond] or "").strip()
        data.append((nombre, precio_unit, precio_mayor, condicion))
    return data


def to_float(value):
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    try:
        return float(value)
    except Exception:
        raise ValueError(f"No se pudo convertir a número: {value!r}")


def prepare(raw_path: Path, out_path: Path):
    rows = read_raw(raw_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "catalogo"
    headers = [
        "sku",
        "nombre",
        "categoria",
        "precio_unit",
        "precio_mayor",
        "umbral_mayor",
        "favorito",
        "superfavorito",
        "visible",
    ]
    ws.append(headers)
    slug_registry: Dict[str, int] = {}
    for nombre, pu, pm, cond in rows:
        base = slugify(nombre)
        sku = unique_slug(base, slug_registry)
        ws.append(
            [
                sku,
                nombre,
                "sin_categoria",
                to_float(pu),
                to_float(pm),
                cond,
                0,
                0,
                1,
            ]
        )
    wb.save(out_path)
    return out_path


def main():
    import argparse

    parser = argparse.ArgumentParser(description="Prepara un Excel crudo al formato estándar de catálogo.")
    parser.add_argument("input", type=Path, help="Excel de entrada con columnas Producto, Precio Unidad, Precio Mayor, Condición")
    parser.add_argument("--out", type=Path, default=Path("catalog_prepared.xlsx"), help="Excel de salida")
    args = parser.parse_args()

    out = prepare(args.input, args.out)
    print(f"Generado {out}")


if __name__ == "__main__":
    main()
