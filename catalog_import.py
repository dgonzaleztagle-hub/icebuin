#!/usr/bin/env python3
"""
Script de importación para catálogo.

Lee un Excel con columnas:
  sku, nombre, categoria, precio_unit, precio_mayor, umbral_mayor, favorito, superfavorito, visible

Reglas:
- Imagen esperada: images_dir/sku.jpg (solo JPG). Si no existe, se reporta.
- Orden por categoría: superfavorito (máx 1) primero, luego favoritos (alfabético), luego resto (alfabético).
- Si un sku desaparece del Excel, se marca visible=False en el catálogo resultante.
- Si precio_mayor está vacío, se considera "sin mayorista" (valor None) y se marca en cambios si aplica.

Outputs (en out_dir):
- updated_catalog.json: catálogo final con orden asignado.
- diff.json: altas, bajas (visibles a False), cambios de campos, imágenes faltantes, warnings.
- summary.txt: resumen humano de lo detectado.
"""
from __future__ import annotations

import argparse
import json
import sys
from dataclasses import dataclass, asdict
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - dependencia no estándar
    sys.stderr.write("Falta dependencia: pip install openpyxl\n")
    raise exc


REQUIRED_COLUMNS = [
    "sku",
    "nombre",
    "categoria",
    "precio_unit",
    "precio_mayor",
    "umbral_mayor",
    "favorito",
    "superfavorito",
    "visible",
]

# Aliases comunes en español para mapear a las columnas requeridas.
ALIASES = {
    "producto": "nombre",
    "condicion": "umbral_mayor",
    "condición": "umbral_mayor",
    "precio unidad": "precio_unit",
    "precio unidad.": "precio_unit",
    "precio mayor": "precio_mayor",
}


def to_bool(value) -> bool:
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return value != 0
    text = str(value).strip().lower()
    return text in {"1", "true", "si", "sí", "yes", "y", "ok", "x"}


def to_number(value) -> Optional[float]:
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        raise ValueError(f"No se pudo convertir a número: {value!r}")


@dataclass
class Product:
    sku: str
    nombre: str
    categoria: str
    precio_unit: float
    precio_mayor: Optional[float]
    umbral_mayor: Optional[str]
    favorito: bool
    superfavorito: bool
    visible: bool
    imagen: str

    def ordering_key(self) -> Tuple[int, str]:
        priority = 0 if self.superfavorito else 1 if self.favorito else 2
        return (priority, self.nombre.lower())


def load_excel(path: Path) -> Dict[str, Product]:
    wb = load_workbook(path)
    sheet = wb.active
    rows = sheet.iter_rows(values_only=True)
    try:
        header = [str(h).strip().lower() for h in next(rows)]
        # Aplicar alias: si viene "condicion" o "producto" etc., los renombramos.
        normalized_header = []
        for name in header:
            normalized_header.append(ALIASES.get(name, name))
        header = normalized_header
    except StopIteration:
        raise ValueError("El archivo está vacío")

    col_index = {name: idx for idx, name in enumerate(header)}
    missing = [c for c in REQUIRED_COLUMNS if c not in col_index]
    if missing:
        raise ValueError(f"Faltan columnas requeridas: {', '.join(missing)}")

    products: Dict[str, Product] = {}
    for raw in rows:
        if raw is None or all(cell is None or str(cell).strip() == "" for cell in raw):
            continue

        def get(col: str):
            idx = col_index[col]
            return raw[idx] if idx < len(raw) else None

        sku = str(get("sku") or "").strip()
        if not sku:
            continue

        precio_unit = to_number(get("precio_unit"))
        if precio_unit is None:
            raise ValueError(f"El producto {sku} no tiene precio_unit válido")

        precio_mayor = to_number(get("precio_mayor"))
        producto = Product(
            sku=sku,
            nombre=str(get("nombre") or "").strip(),
            categoria=str(get("categoria") or "").strip(),
            precio_unit=precio_unit,
            precio_mayor=precio_mayor,  # None => sin mayorista
            umbral_mayor=str(get("umbral_mayor") or "").strip() or None,
            favorito=to_bool(get("favorito")),
            superfavorito=to_bool(get("superfavorito")),
            visible=to_bool(get("visible")),
            imagen=f"{sku}.jpg",
        )
        products[sku] = producto
    return products


def load_json(path: Path) -> Dict[str, Product]:
    if not path.exists():
        return {}
    data = json.loads(path.read_text(encoding="utf-8"))
    result: Dict[str, Product] = {}
    for item in data:
        result[item["sku"]] = Product(
            sku=item["sku"],
            nombre=item["nombre"],
            categoria=item["categoria"],
            precio_unit=float(item["precio_unit"]),
            precio_mayor=item.get("precio_mayor"),
            umbral_mayor=item.get("umbral_mayor"),
            favorito=bool(item.get("favorito", False)),
            superfavorito=bool(item.get("superfavorito", False)),
            visible=bool(item.get("visible", True)),
            imagen=item.get("imagen") or f"{item['sku']}.jpg",
        )
    return result


def validate_superfavorito(products: Dict[str, Product]) -> List[str]:
    warnings = []
    by_cat: Dict[str, List[Product]] = {}
    for p in products.values():
        by_cat.setdefault(p.categoria, []).append(p)
    for cat, items in by_cat.items():
        count = sum(1 for p in items if p.superfavorito)
        if count > 1:
            warnings.append(
                f"Categoría '{cat}' tiene {count} superfavoritos (máx 1)."
            )
    return warnings


def diff_catalog(
    new: Dict[str, Product], old: Dict[str, Product]
) -> Tuple[Dict[str, Product], Dict]:
    updated: Dict[str, Product] = dict(new)
    diff = {"added": [], "removed": [], "changed": []}

    for sku, product in new.items():
        if sku not in old:
            diff["added"].append(sku)
        else:
            changes = {}
            prev = old[sku]
            for field in [
                "nombre",
                "categoria",
                "precio_unit",
                "precio_mayor",
                "umbral_mayor",
                "favorito",
                "superfavorito",
                "visible",
                "imagen",
            ]:
                if getattr(prev, field) != getattr(product, field):
                    changes[field] = {
                        "from": getattr(prev, field),
                        "to": getattr(product, field),
                    }
            if changes:
                diff["changed"].append({"sku": sku, "changes": changes})

    for sku, product in old.items():
        if sku not in new:
            hidden_data = {**asdict(product), "visible": False}
            hidden = Product(**hidden_data)
            updated[sku] = hidden
            diff["removed"].append(sku)

    return updated, diff


def order_by_category(products: Dict[str, Product]) -> Dict[str, List[Product]]:
    grouped: Dict[str, List[Product]] = {}
    for p in products.values():
        grouped.setdefault(p.categoria, []).append(p)
    for cat, items in grouped.items():
        items.sort(key=lambda p: p.ordering_key())
    return grouped


def find_missing_images(products: Dict[str, Product], images_dir: Path) -> List[str]:
    missing = []
    for p in products.values():
        if not (images_dir / p.imagen).exists():
            missing.append(p.sku)
    return missing


def write_outputs(
    out_dir: Path,
    grouped: Dict[str, List[Product]],
    diff: Dict,
    warnings: List[str],
    missing_images: List[str],
) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)
    catalog_payload = []
    for cat, items in grouped.items():
        for order, product in enumerate(items, start=1):
            payload = asdict(product)
            payload["orden_categoria"] = order
            catalog_payload.append(payload)

    (out_dir / "updated_catalog.json").write_text(
        json.dumps(catalog_payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    summary = {
        "timestamp": datetime.utcnow().isoformat() + "Z",
        "added": diff["added"],
        "removed_set_visible_false": diff["removed"],
        "changed": diff["changed"],
        "missing_images": missing_images,
        "warnings": warnings,
    }
    (out_dir / "diff.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    lines = [
        f"Importación: {summary['timestamp']}",
        f"Altas: {len(diff['added'])} -> {', '.join(diff['added']) if diff['added'] else 'ninguna'}",
        f"Bajas (visibles a False): {len(diff['removed'])} -> {', '.join(diff['removed']) if diff['removed'] else 'ninguna'}",
        f"Cambios: {len(diff['changed'])}",
        f"Imágenes faltantes (.jpg): {len(missing_images)} -> {', '.join(missing_images) if missing_images else 'ninguna'}",
    ]
    if warnings:
        lines.append("Warnings:")
        lines.extend(f"- {w}" for w in warnings)
    (out_dir / "summary.txt").write_text("\n".join(lines), encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Importador de catálogo desde Excel.")
    parser.add_argument("excel", type=Path, help="Ruta al archivo .xlsx de entrada")
    parser.add_argument(
        "--current",
        type=Path,
        default=Path("current_catalog.json"),
        help="Catálogo actual en JSON (opcional)",
    )
    parser.add_argument(
        "--images-dir",
        type=Path,
        default=Path("images"),
        help="Carpeta con imágenes sku.jpg",
    )
    parser.add_argument(
        "--out-dir",
        type=Path,
        default=Path("out"),
        help="Carpeta de salida para JSON y reportes",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    try:
        new_products = load_excel(args.excel)
    except Exception as exc:  # pragma: no cover - errores de datos de entrada
        sys.stderr.write(f"Error al leer Excel: {exc}\n")
        return 1

    old_products = load_json(args.current)
    warnings = validate_superfavorito(new_products)
    updated_catalog, diff = diff_catalog(new_products, old_products)
    grouped = order_by_category(updated_catalog)
    missing_images = find_missing_images(updated_catalog, args.images_dir)
    write_outputs(args.out_dir, grouped, diff, warnings, missing_images)

    print("Importación completada. Revisa out/summary.txt y out/diff.json.")
    if warnings:
        print("Warnings:")
        for w in warnings:
            print("-", w)
    return 0


if __name__ == "__main__":
    sys.exit(main())
